"""Normalize creator rows collected from FastMoss web pages or web exports."""
from __future__ import annotations

import csv
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:  # CSV mode still works without openpyxl.
    load_workbook = None


ALIASES = {
    "username": ("username", "unique_id", "creator_unique_id", "handle", "达人账号", "达人用户名", "用户名"),
    "country": ("country", "region", "国家", "地区", "国家/地区"),
    "followers": ("followers", "follower_count", "粉丝", "粉丝数"),
    "email": ("email", "邮箱", "联系邮箱", "商务邮箱"),
    "units_sold": ("units_sold", "sales", "销量", "带货销量"),
    "gmv": ("gmv", "销售额", "带货gmv", "带货销售额"),
    "shop_valid": ("shop_valid", "has_showcase", "showcase", "是否带货", "带货达人"),
    "bio": ("bio", "signature", "简介"),
    "profile_url": ("profile_url", "tiktok_url", "达人主页"),
    "source_url": ("source_url", "fastmoss_url", "详情页"),
}


def _pick(row: dict, field: str, default=""):
    lowered = {str(key).strip().lower(): value for key, value in row.items()}
    for alias in ALIASES[field]:
        if alias.lower() in lowered and lowered[alias.lower()] not in (None, ""):
            return lowered[alias.lower()]
    return default


def _scaled_number(value, *, integer: bool):
    try:
        text = (
            str(value).replace(",", "").replace("$", "")
            .replace("€", "").replace("£", "").strip().lower()
        )
        multiplier = 1
        if text.endswith("k"):
            text, multiplier = text[:-1], 1_000
        elif text.endswith("m"):
            text, multiplier = text[:-1], 1_000_000
        elif integer and text.endswith("b"):
            text, multiplier = text[:-1], 1_000_000_000
        result = float(text) * multiplier
        return int(result) if integer else result
    except (TypeError, ValueError):
        return 0 if integer else 0.0


def _integer(value) -> int:
    return _scaled_number(value, integer=True)


def _number(value) -> float:
    return _scaled_number(value, integer=False)


def _truthy(value) -> bool:
    return str(value).strip().lower() in {"1", "true", "yes", "y", "是", "有效", "有"}


def _file_rows(path: Path) -> list[dict]:
    if path.suffix.lower() == ".csv":
        with path.open(encoding="utf-8-sig", newline="") as handle:
            return list(csv.DictReader(handle))
    if path.suffix.lower() in {".xlsx", ".xlsm"} and load_workbook:
        sheet = load_workbook(path, read_only=True, data_only=True).active
        values = sheet.iter_rows(values_only=True)
        headers = [str(value or "").strip() for value in next(values)]
        return [dict(zip(headers, row)) for row in values]
    raise ValueError(f"unsupported FastMoss export: {path}")


def load_fastmoss_exports(paths: list[str], countries: set[str], max_followers: int) -> list[dict]:
    """Normalize FastMoss web collector/export rows. Filter provenance is proof."""
    candidates: dict[str, dict] = {}
    for name in paths:
        path = Path(name).expanduser()
        for row in _file_rows(path):
            username = str(_pick(row, "username")).strip().lstrip("@").lower()
            country = str(_pick(row, "country")).strip().upper()
            followers = _integer(_pick(row, "followers"))
            if not username or country not in countries or not 0 < followers < max_followers:
                continue
            units_sold = _integer(_pick(row, "units_sold"))
            gmv = _number(_pick(row, "gmv"))
            shop_valid = units_sold > 0 or gmv > 0 or _truthy(_pick(row, "shop_valid"))
            if not shop_valid:
                continue
            candidates[username] = {
                "username": username,
                "country": country,
                "followers": followers,
                "email": str(_pick(row, "email")).strip().lower(),
                "bio": str(_pick(row, "bio")).strip(),
                "fastmoss_units_sold": units_sold,
                "fastmoss_gmv": gmv,
                "shop_valid": shop_valid,
                "shop_proof": "shop_showcase_verified",
                "shop_proof_method": str(
                    row.get("shop_proof_method") or "fastmoss_filtered_web_page"
                ),
                "source": str(row.get("source") or "fastmoss_web_export"),
                "profile_url": str(_pick(row, "profile_url")),
                "source_url": str(_pick(row, "source_url")),
            }
    return list(candidates.values())
